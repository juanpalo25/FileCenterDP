import io
from datetime import date, datetime

import openpyxl

from config import ESTADO_PENDIENTE, ESTADO_FINAL_POR_TIPO
from db import get_conn
from maestros import buscar_producto_por_sku


class PlantillaInvalida(Exception):
    pass


_COLUMNAS_REQUERIDAS = {
    "ODC": ["Marca", "SKU", "Cantidad", "Costo_actualizado"],
    "ODR": ["SKU", "Cantidad"],
    "CDP": ["SKU", "PVP", "Costo"],
}


def parsear_plantilla(tipo: str, archivo_bytes: bytes) -> list[dict]:
    """Lee la plantilla Excel subida por el analista y devuelve la lista de items.
    FDP no tiene columnas obligatorias: se valida solo que el archivo no esté vacío."""
    if tipo == "FDP":
        if not archivo_bytes:
            raise PlantillaInvalida("El archivo está vacío.")
        return []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(archivo_bytes), data_only=True)
    except Exception as exc:
        raise PlantillaInvalida(f"No se pudo leer el archivo Excel: {exc}") from exc
    ws = wb.worksheets[0]

    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    requeridas = _COLUMNAS_REQUERIDAS[tipo]
    faltantes = [c for c in requeridas if c not in headers]
    if faltantes:
        raise PlantillaInvalida(
            f"A la plantilla le faltan las columnas obligatorias: {', '.join(faltantes)}"
        )
    idx = {h: i for i, h in enumerate(headers)}

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["SKU"]] is None:
            continue
        sku = row[idx["SKU"]]
        try:
            sku = int(sku)
        except (TypeError, ValueError):
            raise PlantillaInvalida(f"SKU inválido: {sku!r}")

        item = {"sku": sku}
        if tipo == "ODC":
            marca = row[idx["Marca"]]
            cantidad = row[idx["Cantidad"]]
            costo_actualizado = row[idx["Costo_actualizado"]]
            if marca is None or str(marca).strip() == "":
                raise PlantillaInvalida(f"Falta Marca para el SKU {sku}")
            if cantidad is None:
                raise PlantillaInvalida(f"Falta Cantidad para el SKU {sku}")
            item["marca"] = str(marca).strip()
            item["cantidad"] = cantidad
            item["costo_actualizado"] = costo_actualizado
        elif tipo == "ODR":
            cantidad = row[idx["Cantidad"]]
            if cantidad is None:
                raise PlantillaInvalida(f"Falta Cantidad para el SKU {sku}")
            item["cantidad"] = cantidad
        elif tipo == "CDP":
            pvp = row[idx["PVP"]]
            if pvp is None:
                raise PlantillaInvalida(f"Falta PVP para el SKU {sku}")
            item["pvp"] = pvp
            item["costo"] = row[idx["Costo"]]
        items.append(item)

    if not items:
        raise PlantillaInvalida("La plantilla no tiene ningún renglón cargado.")
    return items


def agrupar_por_marca(items: list[dict]) -> dict[str, list[dict]]:
    """Solo para ODC: agrupa los items por marca, preservando el orden de aparición
    tanto de las marcas como de los items dentro de cada una."""
    grupos: dict[str, list[dict]] = {}
    for item in items:
        grupos.setdefault(item["marca"], []).append(item)
    return grupos


def detectar_diferencias_costo(items: list[dict]) -> list[dict]:
    """Solo para ODC: compara Costo_actualizado contra el Costo Ppal de MaestroDP por SKU."""
    diferencias = []
    for item in items:
        producto = buscar_producto_por_sku(item["sku"])
        costo_maestro = producto["costo_ppal"] if producto else None
        item["costo_maestro"] = costo_maestro
        costo_actualizado = item.get("costo_actualizado")
        if costo_maestro is not None and costo_actualizado is not None and float(costo_maestro) != float(costo_actualizado):
            diferencias.append(
                {
                    "sku": item["sku"],
                    "costo_maestro": costo_maestro,
                    "costo_actualizado": costo_actualizado,
                }
            )
    return diferencias


def crear_solicitud(
    tipo: str,
    comitente: str,
    rubro: str,
    prioridad: str,
    archivo_nombre: str,
    archivo_bytes: bytes,
    items: list[dict],
    creado_por: str,
    fecha_vigencia: str | None = None,
    marca: str | None = None,
) -> int:
    ahora = datetime.now().isoformat(timespec="seconds")
    with get_conn() as conn:
        cur = conn.execute(
            """INSERT INTO solicitudes
               (tipo, comitente, rubro, marca, prioridad, fecha_vigencia, estado, fecha_creacion,
                archivo_origen_nombre, archivo_origen_datos, creado_por)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                tipo,
                comitente,
                rubro,
                marca,
                prioridad,
                fecha_vigencia,
                ESTADO_PENDIENTE,
                ahora,
                archivo_nombre,
                archivo_bytes,
                creado_por,
            ),
        )
        solicitud_id = cur.lastrowid

        for item in items:
            conn.execute(
                """INSERT INTO solicitud_items
                   (solicitud_id, sku, cantidad, costo_actualizado, costo_maestro, pvp, costo)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (
                    solicitud_id,
                    item.get("sku"),
                    item.get("cantidad"),
                    item.get("costo_actualizado"),
                    item.get("costo_maestro"),
                    item.get("pvp"),
                    item.get("costo"),
                ),
            )

        conn.execute(
            """INSERT INTO historial_estados (solicitud_id, estado, fecha, usuario, detalle)
               VALUES (?, ?, ?, ?, ?)""",
            (solicitud_id, ESTADO_PENDIENTE, ahora, creado_por, "Solicitud creada"),
        )
    return solicitud_id


def listar_solicitudes(
    tipo: str | None = None,
    comitente: str | None = None,
    rubro: str | None = None,
    prioridad: str | None = None,
    estado: str | None = None,
) -> list[dict]:
    query = "SELECT * FROM solicitudes WHERE 1=1"
    params = []
    for campo, valor in [
        ("tipo", tipo),
        ("comitente", comitente),
        ("rubro", rubro),
        ("prioridad", prioridad),
        ("estado", estado),
    ]:
        if valor:
            query += f" AND {campo} = ?"
            params.append(valor)
    query += " ORDER BY id DESC"
    with get_conn() as conn:
        rows = conn.execute(query, params).fetchall()
    return [dict(r) for r in rows]


def obtener_solicitud(solicitud_id: int) -> dict | None:
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM solicitudes WHERE id = ?", (solicitud_id,)).fetchone()
        if row is None:
            return None
        items = conn.execute(
            "SELECT * FROM solicitud_items WHERE solicitud_id = ?", (solicitud_id,)
        ).fetchall()
        historial = conn.execute(
            "SELECT * FROM historial_estados WHERE solicitud_id = ? ORDER BY fecha", (solicitud_id,)
        ).fetchall()
    data = dict(row)
    data["items"] = [dict(i) for i in items]
    data["historial"] = [dict(h) for h in historial]
    return data


def actualizar_estado(solicitud_id: int, referencia_externa: str, usuario: str):
    solicitud = obtener_solicitud(solicitud_id)
    if solicitud is None:
        raise ValueError("Solicitud no encontrada")
    nuevo_estado = ESTADO_FINAL_POR_TIPO[solicitud["tipo"]]
    ahora = datetime.now().isoformat(timespec="seconds")
    with get_conn() as conn:
        conn.execute(
            """UPDATE solicitudes
               SET estado = ?, fecha_emision = ?, referencia_externa = ?, actualizado_por = ?
               WHERE id = ?""",
            (nuevo_estado, ahora, referencia_externa, usuario, solicitud_id),
        )
        conn.execute(
            """INSERT INTO historial_estados (solicitud_id, estado, fecha, usuario, detalle)
               VALUES (?, ?, ?, ?, ?)""",
            (solicitud_id, nuevo_estado, ahora, usuario, f"Referencia externa: {referencia_externa}"),
        )
